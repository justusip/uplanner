import datetime
import json
from openpyxl import load_workbook


def add(db: [any], filename: str, terms_to_add: [str]):
    workbook = load_workbook(filename=filename)
    sheet = workbook.active
    c_dict = {}

    i = -1
    header = []
    for raw_row in sheet.iter_rows(values_only=True):
        i += 1
        if i == 0:
            header = [str(o).strip() for o in raw_row]
            continue

        row = {k: str(raw_row[i]).strip() for i, k in enumerate(header)}

        r_term_name = row["TERM"]
        r_course_code = row["COURSE CODE"]
        r_section_code = row["CLASS SECTION"]

        r_start_date = row["START DATE"]
        r_end_date = row["END DATE"]

        r_venue = row["VENUE"]
        r_start_time = row["START TIME"]
        r_end_time = row["END TIME"]
        r_course_title = row["COURSE TITLE"]

        quantified_name = f"{r_course_code}/{r_term_name}"
        if r_term_name not in terms_to_add:
            print(f"Skipping {quantified_name}...")
            continue

        print(f"Processing {quantified_name}...")

        if quantified_name not in c_dict:
            c_dict[quantified_name] = {
                "code": r_course_code,
                "term": r_term_name,
                "title": r_course_title,
                "sections": {}
            }

        if r_section_code not in c_dict[quantified_name]["sections"]:
            c_dict[quantified_name]["sections"][r_section_code] = []

        if r_start_date == "" or r_end_date == "":
            print(f"{quantified_name}: Start date or end date missing")
            continue
        if r_start_time == "" or r_end_time == "":
            print(f"{quantified_name}: Start time or end time missing")
            continue

        start_date = datetime.datetime.strptime(r_start_date, "%Y-%m-%d").date()
        end_date = datetime.datetime.strptime(r_end_date, "%Y-%m-%d").date()

        weekday = None
        for i in range(7):
            if raw_row[header.index("MON") + i] is not None:
                weekday = i
        if weekday is None:
            print(f"{quantified_name}: None of the weekday column is not empty.")
            continue

        start_time = datetime.datetime.strptime(r_start_time, "%H:%M").time()
        end_time = datetime.datetime.strptime(r_end_time, "%H:%M").time()

        for i in range((end_date - start_date).days + 1):
            day = start_date + datetime.timedelta(days=i)
            if day.weekday() == weekday:
                start = datetime.datetime.combine(day, start_time)
                end = datetime.datetime.combine(day, end_time)
                c_dict[quantified_name]["sections"][r_section_code].append({
                    "from": start.isoformat(),
                    "to": end.isoformat(),
                    "venue": r_venue,
                })

        c_dict[quantified_name]["sections"][r_section_code].sort(key=lambda o: o["from"], reverse=False)

    c_list = list(c_dict.values())
    for course in c_list:
        course["sections"] = [
            {
                "sectionName": sectionCode,
                "times": times
            } for sectionCode, times in course["sections"].items()
        ]

    db.extend(c_list)


courses = []

add(courses, "2021-22_class_timetable_sem1.xlsx", ["2021-22 Sem 1"])
add(courses, "2021-22_class_timetable_20220112.xlsx", ["2021-22 Sem 2", "2021-22 Sum Sem"])

courses.sort(key=lambda c: c["code"])

with open("../public/out.json", "w", encoding="utf8") as outFile:
    json.dump(courses, outFile, sort_keys=True, indent=4, ensure_ascii=False)
